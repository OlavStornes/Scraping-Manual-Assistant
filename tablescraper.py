import openpyxl
from openpyxl.styles.borders import Border, Side
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select


class Scrapie():
    def __init__(self, system):
        self.page_number = 1
        self.target_system = system
        self.sheet_current_row = 1
        self.finished = False
        self.elem_id = "GridView1"
        self.db_name = 'Gamelist.xlsx'
        self.system_dropid = "DropSys"
        self.url = "http://www.gamesdatabase.org/list.aspx?manual=1"
        self.js_script = "javascript:__doPostBack('" + \
            self.elem_id + \
            "','Page${0}')"

    def init_browser(self):
        self.browser = webdriver.Chrome()
        self.browser.get(self.url)

    def db_new_sheet(self):
        self.db.create_sheet(self.target_system)

        sheet = self.db[self.target_system]

        # Spacing hack
        sheet.append([
            '______________________Game______________________',
            '___System___',
            '________Publisher________',
            '_______Developer_______',
            '__Category__',
            '__Year__',
        ])

        # Styling
        for cell in sheet[1:1]:
            cell.font = cell.font.copy(bold=True, italic=True)
            cell.border = Border(
                bottom=Side(style='thick'))

        # Column spacing
        for col in sheet.columns:
            column = col[0].column
            length = len(col[0].value)

            sheet.column_dimensions[column].width = length

        self.sheet_current_row += 1
        self.db_save()
        return sheet

    def db_save(self):
        print("Saving..")
        try:
            self.db.save(self.db_name)
        except PermissionError as e:
            print(e)

    def init_db(self):
        self.db = openpyxl.load_workbook(self.db_name)

        try:
            self.sheet = self.db[self.target_system]
        except KeyError as e:
            print(e)
            self.sheet = self.db_new_sheet()

    def setup_browser_table(self):
        selector = Select(self.browser.find_element_by_id(self.system_dropid))
        selector.select_by_visible_text(self.target_system)

        delay = 2  # seconds
        try:
            self.table = WebDriverWait(self.browser, delay).until(
                EC.presence_of_element_located((By.ID, self.elem_id)))
        except TimeoutException:
            print("Loading took too much time! Is it the last page?")
            if '404' in self.browser.title:
                self.finished = True

    def aquire_table(self):
        self.table = self.browser.find_element_by_id("GridView1")

    def iterate_pages(self):
        self.page_number += 1
        self.browser.execute_script(self.js_script.format(self.page_number))

        delay = 2  # seconds
        try:
            self.table = WebDriverWait(self.browser, delay).until(
                EC.presence_of_element_located((By.ID, self.elem_id)))
            print("Page is ready!")
        except TimeoutException:
            print("Loading took too much time! Is it the last page?")
            if '404' in self.browser.title:
                self.finished = True

    def aquire_entries(self):
        entries = self.table.find_elements_by_tag_name('tr')[3:]
        self.aquire_cells(entries)

    def aquire_cells(self, entries):
        for index_row, row in enumerate(entries[:-2]):
            self.write_row_to_database(row)

        print("Page is done!")
        self.db_save()

    def cell_contains_text(self, rawcell):
        if rawcell.text != '':
            return True
        else:
            return False

    def write_row_to_database(self, row):
        rawcell = row.find_elements_by_tag_name('td')

        cells = filter(self.cell_contains_text, rawcell)

        for index_col, cell in enumerate(cells):
            # Row offset - due to header
            # Col offset - arrays start at 1 :^)
            _ = self.sheet.cell(
                column=index_col+1,
                row=self.sheet_current_row,
                value=cell.text
            )

        print("Inserted in row {}: {}".format(
            self.sheet_current_row,
            row.text
        ))
        self.sheet_current_row += 1

    def scrape_all(self):
        self.aquire_table()
        while not self.finished:
            self.aquire_entries()
            self.iterate_pages()

    def run(self):
        self.init_browser()
        self.setup_browser_table()
        self.init_db()
        self.scrape_all()


if __name__ == "__main__":
    scraper = Scrapie('Amstrad CPC')
    scraper.run()
